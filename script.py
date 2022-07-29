from openpyxl import load_workbook
import mysql.connector

# Se carga el archivo 
wb = load_workbook(filename = "datos.xlsx")

# Se extrae la hoja donde estan los datos
ws = wb['Hoja1']

# Se mira la cantidad de personas que existen en la hoja de datos
col_length = len(ws['A'])
numero_personas = col_length - 1

# Se extrae el nombre de los campos
dataNames = []
for names in ws['1']:
    dataNames.append(names.value)

# Se inicializa el diccionario 
dataFinal = {}

# Se empiezan a extraer los datos de cada fila
for persona in range(numero_personas):
    dataValues = []
    for datos in ws[str(persona + 2)]:
        dataValues.append(datos.value)

    # Se crea un diccionario con el nombre de los datos y los datos extraidos
    dataFinal["Persona "+str(persona)] = dict(zip(dataNames,dataValues))

    # Se arreglan los datos de la casilla 'Data', para esto se evaluan resultando un diccionario
    dataTemp = eval(dataFinal["Persona " + str(persona)]['Data'])

    # El campo data ya no necesario por lo que se borra
    del dataFinal["Persona " + str(persona)]['Data']

    # Se actualiza el diccionario con los datos arreglados
    dataFinal["Persona " + str(persona)].update(dataTemp)
 
# Inicio de la coneccion con la base de datos
# Se manejan exepciones en caso de que no se conecte a la base de datos 
try: 

    # Se crea la conexion con la base de datos
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="root",
        port="23306",
        database="punto2"
        )

    # Se crea un cursor para poder ejecutar instruciones SQL    
    mycursor = mydb.cursor()

    # Se crea la forma de la instruccion que se va a ejecutar
    sql = "INSERT INTO datos_xlxs (Cedula, Nombres, Direccion, latitude, longitude, city, description) VALUES (%s,%s,%s,%s,%s,%s,%s)"

    # Se recorre el diccionario con los datos de las personas y se agregan a la base de datos
    for persona in dataFinal:

        # Se convierte los valores de los campos a una tupla
        val = tuple(dataFinal[persona].values())

        # Se ejecuta la instruccion SQL enviando los datos de la tupla anterior
        mycursor.execute(sql, val)
        mydb.commit()
        print(mycursor.rowcount, "Datos insertados.")

    # Cierra la conexion con la base de datos
    mycursor.close()
    mydb.close()

# Exepcion en caso de que no conecte la base de datos
except mysql.connector.errors.InterfaceError:
    print('No se pudo conectar con la base de datos')
    